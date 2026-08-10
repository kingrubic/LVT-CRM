"""Build the browser data bundle from the supplied acceptance backlog workbook."""
import json
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "Backlog_Quan_ly_so_Truong_hoc_VSC.xlsx"

wb = load_workbook(SOURCE, data_only=True)
ws = wb["Backlog"]
headers = [cell.value for cell in ws[1]]
keys = {
    "STT": "no", "Issue ID": "id", "Loại Issue": "type", "Epic": "epic",
    "Module": "module", "Tên tính năng": "name", "Mô tả ngắn": "description",
    "Vai trò sử dụng": "roles", "Nền tảng": "platform", "Ưu tiên": "priority",
    "Mốc triển khai": "milestone", "Phụ thuộc": "depends", "Lưu ý PO và Technical": "notes",
    "Tiêu chí hoàn thành": "acceptance", "Trạng thái": "status", "Người phụ trách": "owner",
    "Sprint": "sprint", "Story Point": "points", "Ngày bắt đầu": "start",
    "Hạn hoàn thành": "deadline", "Tiến độ %": "progress", "QA": "qa", "UAT": "uat",
    "Blocker hoặc ghi chú": "blocker"
}
issues = []
for row in ws.iter_rows(min_row=2, values_only=True):
    item = {keys[h]: v for h, v in zip(headers, row) if h in keys}
    if not item.get("id"):
        continue
    issues.append(item)

inputs_ws = wb["Dau vao can chot"]
inputs = []
for row in inputs_ws.iter_rows(min_row=3, values_only=True):
    if not row[0]:
        continue
    # Skip worksheet column-label row accidentally included below headers.
    if str(row[0]).strip().upper() == "ID":
        continue
    inputs.append({"id": row[0], "module": row[1], "decision": row[2], "owner": row[3],
                   "milestone": row[4], "status": row[5], "impact": row[6], "note": row[7]})

payload = "window.ACCEPTANCE_DATA = " + json.dumps({"issues": issues, "inputs": inputs}, ensure_ascii=False, default=str, separators=(",", ":")) + ";\n"
(ROOT / "data.js").write_text(payload, encoding="utf-8")
print(f"Built {len(issues)} issues and {len(inputs)} decision inputs")
